import os
from django.conf import settings
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from apps.usuarios.decorators import role_required
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from apps.consultas.models import Consulta

# Funciones auxiliares
def si_no(valor):
    return "SI" if valor else "No"
    
def clasificar_imc(imc):
    if imc is None:
        return ""
    if imc < 18.5:
        return "BP"
    elif 18.5 <= imc < 25:
        return "PI"
    elif imc < 30:
        return "SP"
    elif imc < 35:
        return "OG I"
    elif imc < 40:
        return "OG II"
    elif imc < 50:
        return "OG III"
    else: 
        return "OG IV"
    
#función para obtener las consultas por periodos
def obtener_consultas(periodo, anio, mes=None, trimestre=None):

    if periodo == "mensual":
        consultas = Consulta.objects.filter(
            fecha__year = anio,
            fecha__month = int(mes)
        )
    elif periodo == "trimestral":
        TRIMESTRE = {
            1:(1,3),
            2:(4,6),
            3:(7,9),
            4:(10,12),
        }
        mes_inicio, mes_fin = TRIMESTRE[int(trimestre)]

        consultas = Consulta.objects.filter(
            fecha__year = anio,
            fecha__month__gte = mes_inicio,
            fecha__month__lte = mes_fin
        )
    else:
        return None
    
    return consultas.select_related("clave_paciente", "signos_vitales", "categoria_de_padecimiento").order_by("fecha")

#función para limpar R09
def limpiar_r09(ws):
    for row in ws.iter_rows(min_row= 9, max_row= 1000):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None

#función para exportar los datos
def exportar_r09(ws, consultas):
    fila_excel = 9
    consecutivo = 1

    for consulta in consultas:
        paciente = consulta.clave_paciente
        historial = getattr(paciente, "historial", None)
        signos = getattr(consulta, "signos_vitales", None)

        ws[f"A{fila_excel}"] = consecutivo
        ws[f"B{fila_excel}"] = consulta.fecha.strftime("%d/%m/%Y")
        ws[f"C{fila_excel}"] = f"{paciente.nombres} {paciente.apellido_paterno} {paciente.apellido_materno}" if hasattr(paciente, "nombres") and hasattr(paciente, "apellido_paterno") and hasattr(paciente, "apellido_materno") else ""
        ws[f"D{fila_excel}"] = paciente.clave if hasattr(paciente, "clave") else ""
        ws[f"E{fila_excel}"] = str(paciente.carrera_o_puesto)

        ws[f"F{fila_excel}"] = signos.peso if signos else ""
        ws[f"G{fila_excel}"] = signos.talla if signos else ""
        ws[f"H{fila_excel}"] = si_no(historial and historial.usa_cigarro)
        ws[f"I{fila_excel}"] = si_no(historial and historial.ingiere_alcohol)
        ws[f"J{fila_excel}"] = si_no(historial and historial.usa_lentes)
        ws[f"K{fila_excel}"] = (historial.enfermedades_cronicas if historial and historial.enfermedades_cronicas else "")
        ws[f"L{fila_excel}"] = si_no(historial and historial.usa_metodos_anticonceptivos)
        ws[f"M{fila_excel}"] = si_no(historial and historial.es_embarazada)
        ws[f"N{fila_excel}"] = (clasificar_imc(signos.imc) if signos else "")
        ws[f"O{fila_excel}"] = signos.imc if signos else ""
        ws[f"P{fila_excel}"] = si_no(historial and historial.usa_drogas)
        ws[f"Q{fila_excel}"] = f"=F{fila_excel}/(G{fila_excel}*G{fila_excel})"
        ws[f"R{fila_excel}"] = (consulta.categoria_de_padecimiento.padecimiento if consulta.categoria_de_padecimiento else "")
        consecutivo += 1
        fila_excel += 1

def actualizar_r10(ws):

    datos = {
        11:(0, 0),
        12:(0, 0),
        13:(0, 0),
        14:(0, 0),
        15:(0, 0),
        16:(0, 0),
    }

    for fila, valores in datos.items():
        matricula, expediente = valores

        ws[f"D{fila}"] = matricula
        ws[f"E{fila}"] = expediente
    
    

#Funció principal para descargar el reporte
@login_required
@role_required(["medico"])
def descargar_reporte_r09(request):
    
    periodo = request.GET.get("periodo")
    anio = int(request.GET.get("anio"))
    mes = request.GET.get("mes")
    trimestre = request.GET.get("trimestre")

    consultas = obtener_consultas(periodo, anio, mes, trimestre)

    if consultas is None:
        return HttpResponse("Periodo no válido")
    
    ruta = settings.RUTA_PLANTILLA_R09
    
    if not os.path.exists(ruta):
        return HttpResponse("Archivo de plantilla R09 no encontrado")
    try:
        wb = load_workbook(ruta)
    except Exception as e:
        return HttpResponse(f"Error al cargar la plantilla: {str(e)}")
    
    ws_r09 = wb["R09"]
    ws_r10 = wb["R10"]
    limpiar_r09(ws_r09)
    exportar_r09(ws_r09, consultas)
    actualizar_r10(ws_r10)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="R09.xlsx"'

    wb.save(response)
    
    return response